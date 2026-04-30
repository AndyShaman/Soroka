from pathlib import Path
from openpyxl import load_workbook


def extract_xlsx(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)
