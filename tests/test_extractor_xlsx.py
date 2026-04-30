# tests/test_extractor_xlsx.py
from pathlib import Path
from openpyxl import Workbook
from src.adapters.extractors.xlsx import extract_xlsx


def test_extract_xlsx(tmp_path):
    p = tmp_path / "x.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Имя"
    ws["B1"] = "Возраст"
    ws["A2"] = "Андрей"
    ws["B2"] = 35
    wb.save(p)
    text = extract_xlsx(p)
    assert "Имя" in text
    assert "Андрей" in text
    assert "35" in text
